import openpyxl
from openpyxl import Workbook


def Excel1():
    workbook = openpyxl.load_workbook ('C:/PythonPractice/Excel/Python_Demo.xlsx')
    sheet = workbook['Sheet3']
    print (sheet.cell (3, 1).value)
    workbook.close ()


# Excel1()

def Excel2():
    workbook = openpyxl.load_workbook ('C:/PythonPractice/Excel/Python_Demo.xlsx')
    sheet = workbook['Sheet2']
    print (sheet.cell (3, 1).value)
    workbook.close ()


# Excel2()

def Excel3():
    workbook = openpyxl.load_workbook ('C:/PythonPractice/Excel/Python_Demo.xlsx')
    sheet = workbook['Sheet1']
    print (sheet.cell (2, 2).value)
    workbook.close ()


# Excel3()

def Excel4():
    workbook = openpyxl.load_workbook ("C:/PythonPractice/Excel/Python_Demo3.xlsx")
    book = Workbook ()
    sheet = book.active
    rows = (
        (88, 46, 57),
        (89, 38, 12),
        (23, 59, 78),
        (56, 21, 98),
        (24, 18, 43),
        (34, 15, "Sai")
    )

    for row in rows:
        sheet.append (row)

    book.save ('C:/PythonPractice/Excel/Python_Demo3.xlsx')
    book.close ()

    # workbook=Workbook()
    # sheet = workbook.create_sheet("mySheet",0)
    # cell=sheet.cell(1,1)
    # cell.value="Hi"
    # workbook.save(fileName)
    # workbook.close ()


# Excel4 ()

def Excel5():
    workbook = openpyxl.load_workbook('C:/PythonPractice/Excel/Python_Demo3.xlsx')
    sheet = workbook["Sheet"]
    print(sheet.cell(2,1).value)

Excel5()
